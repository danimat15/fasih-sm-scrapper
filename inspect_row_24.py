import openpyxl
import os

def main():
    excel_path = os.path.join("research", "fasih-dashboard-se2026", "Monev Pendataan SE2026 13 Juli 2026.xlsx")
    if not os.path.exists(excel_path):
        print("Monev file not found!")
        return

    wb = openpyxl.load_workbook(excel_path, data_only=False)
    ws = wb["Kecamatan"]
    r = 24
    print("Kecamatan Sheet Row 24 Formulas:")
    for c in range(1, 39):
        val = ws.cell(row=r, column=c).value
        print(f"Col {c} ({openpyxl.utils.get_column_letter(c)}): {val}")

if __name__ == "__main__":
    main()
