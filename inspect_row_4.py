import openpyxl
import os
import sys

def main():
    excel_path = os.path.join("research", "fasih-dashboard-se2026", "Monev Pendataan SE2026 13 Juli 2026.xlsx")
    out_path = "row_4_output_utf8.txt"
    
    if not os.path.exists(excel_path):
        print("Monev file not found!")
        return

    wb = openpyxl.load_workbook(excel_path)
    ws = wb["13 Juli"]
    
    with open(out_path, "w", encoding="utf-8") as f:
        sys.stdout = f
        for col in range(1, 42):
            r1 = ws.cell(row=1, column=col).value
            r2 = ws.cell(row=2, column=col).value
            r3 = ws.cell(row=3, column=col).value
            r4 = ws.cell(row=4, column=col).value
            
            print(f"Col {col} ({openpyxl.utils.get_column_letter(col)}):")
            print(f"  H1: {r1}")
            print(f"  H2: {r2}")
            print(f"  H3: {r3}")
            print(f"  Val/Formula in Row 4: {r4}")
            
    sys.stdout = sys.__stdout__
    print("Done writing row_4_output_utf8.txt")

if __name__ == "__main__":
    main()
