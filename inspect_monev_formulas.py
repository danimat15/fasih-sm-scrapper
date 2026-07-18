import openpyxl
import os
import sys

def main():
    excel_path = os.path.join("research", "fasih-dashboard-se2026", "Monev Pendataan SE2026 13 Juli 2026.xlsx")
    out_path = "inspect_formulas_output_utf8.txt"
    
    if not os.path.exists(excel_path):
        print("Monev file not found!")
        return

    wb = openpyxl.load_workbook(excel_path, data_only=False)
    
    with open(out_path, "w", encoding="utf-8") as f:
        sys.stdout = f
        for sheet_name in wb.sheetnames:
            print("="*60)
            print(f"SHEET: {sheet_name}")
            print("="*60)
            ws = wb[sheet_name]
            
            # Print first 9 rows with cell coordinates, values, and styles (if formulas)
            for r in range(1, 10):
                row_vals = []
                for c in range(1, 42):
                    cell = ws.cell(row=r, column=c)
                    val = cell.value
                    if val is not None:
                        row_vals.append(f"{cell.coordinate}: {val}")
                if row_vals:
                    print(f"Row {r}: {', '.join(row_vals[:12])}")
                    if len(row_vals) > 12:
                        print(f"       ... and {len(row_vals) - 12} more columns")
    
    sys.stdout = sys.__stdout__
    print("Done writing inspect_formulas_output_utf8.txt")

if __name__ == "__main__":
    main()
