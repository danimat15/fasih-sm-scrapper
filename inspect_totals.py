import openpyxl
import os

def main():
    excel_path = os.path.join("research", "fasih-dashboard-se2026", "Monev Pendataan SE2026 13 Juli 2026.xlsx")
    if not os.path.exists(excel_path):
        print("Monev file not found!")
        return

    wb = openpyxl.load_workbook(excel_path, data_only=True)
    for name in wb.sheetnames:
        ws = wb[name]
        print(f"\nSheet: {name} (Total rows: {ws.max_row})")
        # Print the last 5 rows
        for r in range(max(1, ws.max_row - 4), ws.max_row + 1):
            vals = [ws.cell(row=r, column=c).value for c in range(1, 10)]
            print(f"  Row {r}: {vals}")

if __name__ == "__main__":
    main()
