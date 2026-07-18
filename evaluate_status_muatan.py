import openpyxl
import os

def main():
    excel_path = os.path.join("research", "fasih-dashboard-se2026", "Monev Pendataan SE2026 13 Juli 2026.xlsx")
    if not os.path.exists(excel_path):
        print("Monev file not found!")
        return

    wb = openpyxl.load_workbook(excel_path, data_only=False)
    ws = wb["Status Muatan"]
    
    rows_data = []
    # Data starts at row 4
    for r in range(4, 135):
        no = ws.cell(row=r, column=1).value
        name = ws.cell(row=r, column=2).value
        kec = ws.cell(row=r, column=5).value
        
        # Kel td (O, col 15), Kel baru (P, col 16)
        kel_td = ws.cell(row=r, column=15).value or 0
        kel_b = ws.cell(row=r, column=16).value or 0
        
        # Ush td (R, col 18), Ush baru (S, col 19)
        ush_td = ws.cell(row=r, column=18).value or 0
        ush_b = ws.cell(row=r, column=19).value or 0
        
        # Calculate
        kel_sel = kel_b - kel_td
        ush_sel = ush_b - ush_td
        tot_td = kel_td + ush_td
        tot_b = kel_b + ush_b
        tot_sel = tot_b - tot_td
        
        rows_data.append({
            'no': no,
            'name': name,
            'kec': kec,
            'kel_td': kel_td,
            'kel_b': kel_b,
            'kel_sel': kel_sel,
            'ush_td': ush_td,
            'ush_b': ush_b,
            'ush_sel': ush_sel,
            'tot_td': tot_td,
            'tot_b': tot_b,
            'tot_sel': tot_sel
        })
        
    # Print the first 30 rows
    print("=== FIRST 30 ROWS OF STATUS MUATAN ===")
    for i, row in enumerate(rows_data[:30]):
        print(f"{i+1:2d} | {row['name']:<30} | {row['kec']:<30} | KelSel: {row['kel_sel']:4d} | UshSel: {row['ush_sel']:4d} | TotSel: {row['tot_sel']:4d} | TotBaru: {row['tot_b']:4d} | TotTD: {row['tot_td']:4d}")
        
    print("\n=== LAST 15 ROWS OF STATUS MUATAN ===")
    for i, row in enumerate(rows_data[-15:]):
        print(f"{len(rows_data)-14+i:2d} | {row['name']:<30} | {row['kec']:<30} | KelSel: {row['kel_sel']:4d} | UshSel: {row['ush_sel']:4d} | TotSel: {row['tot_sel']:4d} | TotBaru: {row['tot_b']:4d} | TotTD: {row['tot_td']:4d}")

if __name__ == "__main__":
    main()
