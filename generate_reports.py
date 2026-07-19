import os
import shutil
import datetime
import pandas as pd
import openpyxl
from openpyxl.styles import Font, PatternFill, Border, Alignment, Side

# Define WITA timezone (UTC+8)
wita_tz = datetime.timezone(datetime.timedelta(hours=8))

def get_initial_start_time(wita_tz):
    default_time = datetime.datetime.now(wita_tz)
    txt_path = os.path.join("dashboard", "public", "last_updated.txt")
    if os.path.exists(txt_path):
        try:
            with open(txt_path, "r", encoding="utf-8") as f:
                content = f.read().strip()
            if " pukul " in content:
                date_part, time_part = content.split(" pukul ")
                time_part = time_part.replace(" WITA", "").strip().replace(":", ".")
                day_str, month_str, year_str = date_part.split(" ")
                day = int(day_str)
                year = int(year_str)
                months = {
                    "Januari": 1, "Februari": 2, "Maret": 3, "April": 4, "Mei": 5, "Juni": 6,
                    "Juli": 7, "Agustus": 8, "September": 9, "Oktober": 10, "November": 11, "Desember": 12
                }
                month = months.get(month_str, 7)
                hour_str, minute_str = time_part.split(".")
                hour = int(hour_str)
                minute = int(minute_str)
                return datetime.datetime(year, month, day, hour, minute, tzinfo=wita_tz)
        except Exception as e:
            print(f"Error parsing last_updated.txt for START_TIME: {e}")
    return default_time

START_TIME = get_initial_start_time(wita_tz)

def get_current_date_str():
    return f"{START_TIME.day} Juli"

def get_current_timestamp_str():
    months = {
        1: "Januari", 2: "Februari", 3: "Maret", 4: "April", 5: "Mei", 6: "Juni",
        7: "Juli", 8: "Agustus", 9: "September", 10: "Oktober", 11: "November", 12: "Desember"
    }
    day = START_TIME.day
    month_name = months[START_TIME.month]
    year = START_TIME.year
    hour_minute = START_TIME.strftime("%H:%M")
    return f"Kamis, {day} {month_name} {year}", f"{hour_minute} WITA"

def get_last_updated_timestamp(public_dir):
    txt_path = os.path.join(public_dir, "last_updated.txt")
    if os.path.exists(txt_path):
        try:
            with open(txt_path, "r", encoding="utf-8") as f:
                content = f.read().strip()
            if " pukul " in content:
                parts = content.split(" pukul ")
                return parts[0], parts[1]
            return content, ""
        except Exception as e:
            print(f"Error reading last_updated.txt: {e}")
    return get_current_timestamp_str()

def copy_cell_style(src_cell, dest_cell):
    if src_cell.has_style:
        dest_cell.font = Font(
            name=src_cell.font.name,
            size=src_cell.font.size,
            bold=src_cell.font.bold,
            italic=src_cell.font.italic,
            color=src_cell.font.color,
            underline=src_cell.font.underline
        )
        dest_cell.fill = PatternFill(
            fill_type=src_cell.fill.fill_type,
            start_color=src_cell.fill.start_color,
            end_color=src_cell.fill.end_color
        )
        dest_cell.border = Border(
            left=src_cell.border.left,
            right=src_cell.border.right,
            top=src_cell.border.top,
            bottom=src_cell.border.bottom
        )
        dest_cell.alignment = Alignment(
            horizontal=src_cell.alignment.horizontal,
            vertical=src_cell.alignment.vertical,
            wrap_text=src_cell.alignment.wrap_text
        )
        dest_cell.number_format = src_cell.number_format

def calculate_target_info():
    """Calculate daily target info matching frontend logic."""
    import math
    start_date = datetime.datetime(2026, 6, 15, tzinfo=wita_tz)
    now = datetime.datetime.now(wita_tz)
    
    start_day = datetime.datetime(start_date.year, start_date.month, start_date.day, tzinfo=wita_tz)
    current_day = datetime.datetime(now.year, now.month, now.day, tzinfo=wita_tz)
    
    diff_days = (current_day - start_day).days + 1
    
    elapsed_days = diff_days
    if now.hour < 12:
        elapsed_days = diff_days - 1
    elapsed_days = max(0, elapsed_days)
    
    daily_target = 1.67
    cumulative_target = elapsed_days * daily_target
    
    return {
        "daily_target": daily_target,
        "elapsed_days": elapsed_days,
        "cumulative_target": cumulative_target,
    }

def get_realisasi_color(realisasi_pct, cumulative_target):
    """Return (font, fill) tuple based on realisasi vs target thresholds.
    Green: above cumulative target
    Amber/Yellow: between 50% of target and target
    Red: below 50% of target
    """
    green_fill = PatternFill(fill_type="solid", start_color="C6EFCE", end_color="C6EFCE")
    green_font = Font(color="006100", bold=True)
    amber_fill = PatternFill(fill_type="solid", start_color="FFEB9C", end_color="FFEB9C")
    amber_font = Font(color="9C6500", bold=True)
    red_fill = PatternFill(fill_type="solid", start_color="FFC7CE", end_color="FFC7CE")
    red_font = Font(color="9C0006", bold=True)
    
    half_target = 0.5 * cumulative_target
    
    if realisasi_pct >= cumulative_target:
        return green_font, green_fill
    elif realisasi_pct < half_target:
        return red_font, red_fill
    else:
        return amber_font, amber_fill

def format_kec_name(name):
    if not name or pd.isna(name):
        return "-"
    cleaned = str(name).replace(r"/\(\d+\)/g", "").strip()
    # remove leading code like (040)
    if cleaned.startswith("(") and ")" in cleaned:
        cleaned = cleaned.split(")", 1)[1].strip()
    return cleaned.title()

def generate_report_1(public_dir):
    print("Generating Report 1 (Dashboard Leaderboards)...")
    
    hour = START_TIME.hour
    report_type = "pagi" if hour < 13 else "sore"
    scraped_file = "dashboard_scraped_data.csv"
    morning_file = "dashboard_scraped_data_morning.csv"
    morning_prev_file = "dashboard_scraped_data_morning_prev.csv"
    evening_file = "dashboard_scraped_data_evening.csv"
    
    if hour < 13:
        # Report Pagi: current morning vs previous morning
        if os.path.exists(morning_file):
            scraped_file = morning_file
        baseline_file = morning_prev_file
        if not os.path.exists(baseline_file):
            baseline_file = scraped_file
    else:
        # Report Sore: current evening vs morning of same day
        if os.path.exists(evening_file):
            scraped_file = evening_file
        baseline_file = morning_file
        if not os.path.exists(baseline_file):
            baseline_file = scraped_file

    if not os.path.exists(scraped_file):
        print(f"Error: {scraped_file} not found. Skipping Report 1.")
        return
        
    df_curr = pd.read_csv(scraped_file)
    df_curr = df_curr[df_curr["Category"].str.lower() == "pengawas"]
    
    df_base = pd.read_csv(baseline_file)
    df_base = df_base[df_base["Category"].str.lower() == "pengawas"]
    
    # 1. Monitoring per Kecamatan
    # Group current and baseline
    def get_kec_totals(df):
        grouped = df.groupby("nama_kec").agg({
            "OPEN": "sum",
            "DRAFT": "sum",
            "SUBMITTED BY Pencacah": "sum",
            "SUBMITTED RESPONDENT": "sum",
            "REJECTED BY Pengawas": "sum",
            "REJECTED BY Admin Kabupaten": "sum",
            "APPROVED BY Pengawas": "sum",
            "COMPLETED BY Admin Kabupaten": "sum",
            "EDITED BY Admin Kabupaten": "sum",
            "REVOKED BY Pengawas": "sum",
            "SLS Code": "count"
        }).reset_index()
        return grouped
        
    g_curr = get_kec_totals(df_curr)
    g_base = get_kec_totals(df_base)
    
    # Merge current and baseline to find progress harian
    merged_kec = pd.merge(g_curr, g_base, on="nama_kec", how="left", suffixes=("", "_base"))
    
    # Calculate target, open, draft, submitted, rejected, approved, revoke, realisasi
    for suffix in ["", "_base"]:
        sub = "SUBMITTED BY Pencacah" + suffix
        sub_resp = "SUBMITTED RESPONDENT" + suffix
        rej = "REJECTED BY Pengawas" + suffix
        rej_adm = "REJECTED BY Admin Kabupaten" + suffix
        app = "APPROVED BY Pengawas" + suffix
        comp = "COMPLETED BY Admin Kabupaten" + suffix
        edit = "EDITED BY Admin Kabupaten" + suffix
        rev = "REVOKED BY Pengawas" + suffix
        op = "OPEN" + suffix
        dr = "DRAFT" + suffix
        
        merged_kec["Target" + suffix] = merged_kec[op] + merged_kec[dr] + merged_kec[sub] + merged_kec[sub_resp] + merged_kec[rej] + merged_kec[rej_adm] + merged_kec[app] + merged_kec[comp] + merged_kec[edit] + merged_kec[rev]
        merged_kec["Realisasi" + suffix] = merged_kec[sub] + merged_kec[sub_resp] + merged_kec[rej] + merged_kec[rej_adm] + merged_kec[app] + merged_kec[comp] + merged_kec[edit] + merged_kec[rev]
        
    merged_kec["Progres_Harian"] = merged_kec["Realisasi"] - merged_kec["Realisasi_base"].fillna(0)
    merged_kec["Progres_Harian"] = merged_kec["Progres_Harian"].apply(lambda x: max(0, int(x)))
    
    merged_kec["Open_Jml"] = merged_kec["OPEN"]
    merged_kec["Draft_Jml"] = merged_kec["DRAFT"]
    merged_kec["Submitted_Jml"] = merged_kec["SUBMITTED BY Pencacah"] + merged_kec["SUBMITTED RESPONDENT"]
    merged_kec["Rejected_Jml"] = merged_kec["REJECTED BY Pengawas"] + merged_kec["REJECTED BY Admin Kabupaten"]
    merged_kec["Approved_Jml"] = merged_kec["APPROVED BY Pengawas"] + merged_kec["COMPLETED BY Admin Kabupaten"] + merged_kec["EDITED BY Admin Kabupaten"]
    merged_kec["Revoke_Jml"] = merged_kec["REVOKED BY Pengawas"]
    merged_kec["Realisasi_Jml"] = merged_kec["Realisasi"]
    
    def join_unique_kec(series):
        unique_vals = [format_kec_name(val) for val in series.dropna().unique() if val != "-"]
        return ", ".join(sorted(list(set(unique_vals))))

    def join_unique_koseka(series):
        unique_vals = [str(val).strip() for val in series.dropna().unique() if str(val).strip() != "-"]
        return ", ".join(sorted(list(set(unique_vals))))

    # 2. Leaderboards PPL
    # PPL is Category = 'Pencacah'
    df_ppl = pd.read_csv(scraped_file)
    df_ppl = df_ppl[df_ppl["Category"].str.lower() == "pencacah"]
    
    ppl_grouped = df_ppl.groupby("nama_petugas").agg({
        "nama_kec": join_unique_kec,
        "koseka": join_unique_koseka,
        "OPEN": "sum",
        "DRAFT": "sum",
        "SUBMITTED BY Pencacah": "sum",
        "SUBMITTED RESPONDENT": "sum",
        "REJECTED BY Pengawas": "sum",
        "REJECTED BY Admin Kabupaten": "sum",
        "APPROVED BY Pengawas": "sum",
        "COMPLETED BY Admin Kabupaten": "sum",
        "EDITED BY Admin Kabupaten": "sum",
        "REVOKED BY Pengawas": "sum"
    }).reset_index()
    
    ppl_grouped["Target"] = ppl_grouped["OPEN"] + ppl_grouped["DRAFT"] + ppl_grouped["SUBMITTED BY Pencacah"] + ppl_grouped["SUBMITTED RESPONDENT"] + ppl_grouped["REJECTED BY Pengawas"] + ppl_grouped["REJECTED BY Admin Kabupaten"] + ppl_grouped["APPROVED BY Pengawas"] + ppl_grouped["COMPLETED BY Admin Kabupaten"] + ppl_grouped["EDITED BY Admin Kabupaten"] + ppl_grouped["REVOKED BY Pengawas"]
    ppl_grouped["Open"] = ppl_grouped["OPEN"]
    ppl_grouped["Draft"] = ppl_grouped["DRAFT"]
    ppl_grouped["Submit"] = ppl_grouped["SUBMITTED BY Pencacah"] + ppl_grouped["SUBMITTED RESPONDENT"]
    ppl_grouped["Reject"] = ppl_grouped["REJECTED BY Pengawas"] + ppl_grouped["REJECTED BY Admin Kabupaten"]
    ppl_grouped["Approved"] = ppl_grouped["APPROVED BY Pengawas"] + ppl_grouped["COMPLETED BY Admin Kabupaten"] + ppl_grouped["EDITED BY Admin Kabupaten"]
    ppl_grouped["Progres"] = ppl_grouped["REVOKED BY Pengawas"]
    ppl_grouped["Realisasi"] = ppl_grouped["Submit"] + ppl_grouped["Reject"] + ppl_grouped["Approved"] + ppl_grouped["Progres"]
    ppl_grouped["Realisasi_Pct"] = ppl_grouped.apply(lambda r: (r["Realisasi"] / r["Target"] * 100) if r["Target"] > 0 else 0, axis=1)
    
    # 3. PML Leaderboard
    df_pml = pd.read_csv(scraped_file)
    df_pml = df_pml[df_pml["Category"].str.lower() == "pengawas"]
    
    pml_grouped = df_pml.groupby("nama_petugas").agg({
        "nama_kec": join_unique_kec,
        "koseka": join_unique_koseka,
        "OPEN": "sum",
        "DRAFT": "sum",
        "SUBMITTED BY Pencacah": "sum",
        "SUBMITTED RESPONDENT": "sum",
        "REJECTED BY Pengawas": "sum",
        "REJECTED BY Admin Kabupaten": "sum",
        "APPROVED BY Pengawas": "sum",
        "COMPLETED BY Admin Kabupaten": "sum",
        "EDITED BY Admin Kabupaten": "sum",
        "REVOKED BY Pengawas": "sum"
    }).reset_index()
    
    pml_grouped["Target"] = pml_grouped["OPEN"] + pml_grouped["DRAFT"] + pml_grouped["SUBMITTED BY Pencacah"] + pml_grouped["SUBMITTED RESPONDENT"] + pml_grouped["REJECTED BY Pengawas"] + pml_grouped["REJECTED BY Admin Kabupaten"] + pml_grouped["APPROVED BY Pengawas"] + pml_grouped["COMPLETED BY Admin Kabupaten"] + pml_grouped["EDITED BY Admin Kabupaten"] + pml_grouped["REVOKED BY Pengawas"]
    pml_grouped["Open"] = pml_grouped["OPEN"]
    pml_grouped["Draft"] = pml_grouped["DRAFT"]
    pml_grouped["Submit"] = pml_grouped["SUBMITTED BY Pencacah"] + pml_grouped["SUBMITTED RESPONDENT"]
    pml_grouped["Reject"] = pml_grouped["REJECTED BY Pengawas"] + pml_grouped["REJECTED BY Admin Kabupaten"]
    pml_grouped["Approved"] = pml_grouped["APPROVED BY Pengawas"] + pml_grouped["COMPLETED BY Admin Kabupaten"] + pml_grouped["EDITED BY Admin Kabupaten"]
    pml_grouped["Revoke"] = pml_grouped["REVOKED BY Pengawas"]
    pml_grouped["Realisasi"] = pml_grouped["Target"] - pml_grouped["Open"] - pml_grouped["Draft"] - pml_grouped["Submit"]
    pml_grouped["Realisasi_Pct"] = pml_grouped.apply(lambda r: (r["Realisasi"] / r["Target"] * 100) if r["Target"] > 0 else 0, axis=1)
    
    # 4. Kecamatan Leaderboard
    kec_lead = merged_kec.copy()
    kec_lead["Realisasi_Pct"] = kec_lead.apply(lambda r: (r["Realisasi"] / r["Target"] * 100) if r["Target"] > 0 else 0, axis=1)
    kec_lead["Realisasi_Pct_base"] = kec_lead.apply(lambda r: (r["Realisasi_base"] / r["Target_base"] * 100) if r["Target_base"] > 0 else 0, axis=1)
    
    # Sort and rank current
    kec_lead = kec_lead.sort_values(by="Realisasi_Pct", ascending=False).reset_index(drop=True)
    kec_lead["Rank"] = kec_lead.index + 1
    
    # Sort and rank base
    kec_base_sorted = kec_lead.sort_values(by="Realisasi_Pct_base", ascending=False).reset_index(drop=True)
    kec_base_sorted["Rank_base"] = kec_base_sorted.index + 1
    
    # Merge ranks
    kec_lead = pd.merge(kec_lead, kec_base_sorted[["nama_kec", "Rank_base"]], on="nama_kec", how="left")
    kec_lead["Rank_Change"] = kec_lead["Rank_base"].fillna(15) - kec_lead["Rank"]
    
    # Calculate target info
    target_info = calculate_target_info()
    daily_target_pct = target_info["daily_target"]
    elapsed_days = target_info["elapsed_days"]
    cumulative_target_pct = target_info["cumulative_target"]
    
    # Save Report 1 as Excel
    wb = openpyxl.Workbook()
    # Sheet 1: Monitoring Kecamatan
    ws1 = wb.active
    ws1.title = "Monitoring Kecamatan"
    
    # Title
    date_title, time_title = get_last_updated_timestamp(public_dir)
    ws1.merge_cells("A1:K1")
    ws1["A1"] = "MONITORING EVALUASI PROGRES PENDATAAN LAPANGAN SENSUS EKONOMI 2026"
    ws1["A1"].font = Font(name="Calibri", size=14, bold=True)
    ws1["A1"].alignment = Alignment(horizontal="center")
    
    ws1["A3"] = f"Hari/Tanggal : {date_title}"
    ws1["A4"] = f"Pukul        : {time_title}"
    
    # Target info rows
    target_label_font = Font(name="Calibri", size=10, bold=True, color="9C6500")
    target_value_font = Font(name="Calibri", size=10, bold=True, color="E26B0A")
    ws1["D3"] = "Target Harian:"
    ws1["D3"].font = target_label_font
    ws1["E3"] = f"{daily_target_pct}% per hari"
    ws1["E3"].font = target_value_font
    ws1["D4"] = f"Target Hari Ini (H-{elapsed_days}):"
    ws1["D4"].font = target_label_font
    ws1["E4"] = f"{cumulative_target_pct:.2f}%"
    ws1["E4"].font = target_value_font
    
    # Legend for coloring
    ws1["G3"] = "Keterangan Warna Realisasi:"
    ws1["G3"].font = Font(name="Calibri", size=9, bold=True)
    ws1["G4"] = f"Hijau >= {cumulative_target_pct:.2f}%"
    ws1["G4"].font = Font(name="Calibri", size=9, color="006100", bold=True)
    ws1["G4"].fill = PatternFill(fill_type="solid", start_color="C6EFCE", end_color="C6EFCE")
    ws1["H4"] = f"Kuning {cumulative_target_pct*0.5:.2f}% - {cumulative_target_pct:.2f}%"
    ws1["H4"].font = Font(name="Calibri", size=9, color="9C6500", bold=True)
    ws1["H4"].fill = PatternFill(fill_type="solid", start_color="FFEB9C", end_color="FFEB9C")
    ws1["I4"] = f"Merah < {cumulative_target_pct*0.5:.2f}%"
    ws1["I4"].font = Font(name="Calibri", size=9, color="9C0006", bold=True)
    ws1["I4"].fill = PatternFill(fill_type="solid", start_color="FFC7CE", end_color="FFC7CE")
    
    headers = [
        "KECAMATAN", "TARGET", "JUMLAH SLS", 
        "OPEN JML", "OPEN %", "DRAFT JML", "DRAFT %", 
        "SUBMITTED JML", "SUBMITTED %", "REJECTED JML", "REJECTED %",
        "APPROVED JML", "APPROVED %", "REVOKE JML", "REVOKE %",
        "REALISASI JML", "REALISASI %", "PROGRES HARIAN JML", "PROGRES HARIAN %"
    ]
    ws1.append([]) # row 5
    ws1.append(headers) # row 6
    
    # Style headers
    header_fill = PatternFill(fill_type="solid", start_color="E26B0A", end_color="E26B0A")
    header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    for col in range(1, len(headers) + 1):
        cell = ws1.cell(row=6, column=col)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        
    # Populate Kecamatan data
    row_num = 7
    for idx, row in merged_kec.iterrows():
        kec_name = format_kec_name(row["nama_kec"])
        target = row["Target"]
        sls = row["SLS Code"]
        
        open_val = row["Open_Jml"]
        draft_val = row["Draft_Jml"]
        submit_val = row["Submitted_Jml"]
        reject_val = row["Rejected_Jml"]
        approve_val = row["Approved_Jml"]
        revoke_val = row["Revoke_Jml"]
        realisasi = row["Realisasi_Jml"]
        prog_harian = row["Progres_Harian"]
        
        def pct_formula(c_letter, t_letter, r_idx):
            return f"={c_letter}{r_idx}/{t_letter}{r_idx}"
            
        r = row_num
        ws1.cell(row=r, column=1, value=kec_name)
        ws1.cell(row=r, column=2, value=target)
        ws1.cell(row=r, column=3, value=sls)
        
        ws1.cell(row=r, column=4, value=open_val)
        ws1.cell(row=r, column=5, value=pct_formula("D", "B", r))
        
        ws1.cell(row=r, column=6, value=draft_val)
        ws1.cell(row=r, column=7, value=pct_formula("F", "B", r))
        
        ws1.cell(row=r, column=8, value=submit_val)
        ws1.cell(row=r, column=9, value=pct_formula("H", "B", r))
        
        ws1.cell(row=r, column=10, value=reject_val)
        ws1.cell(row=r, column=11, value=pct_formula("J", "B", r))
        
        ws1.cell(row=r, column=12, value=approve_val)
        ws1.cell(row=r, column=13, value=pct_formula("L", "B", r))
        
        ws1.cell(row=r, column=14, value=revoke_val)
        ws1.cell(row=r, column=15, value=pct_formula("N", "B", r))
        
        ws1.cell(row=r, column=16, value=realisasi)
        ws1.cell(row=r, column=17, value=pct_formula("P", "B", r))
        
        ws1.cell(row=r, column=18, value=prog_harian)
        ws1.cell(row=r, column=19, value=pct_formula("R", "B", r))
        
        # Formats
        for col_idx in [5, 7, 9, 11, 13, 15, 17, 19]:
            ws1.cell(row=r, column=col_idx).number_format = '0.00%'
        
        # Conditional coloring for Realisasi % (column 17)
        realisasi_pct_val = (realisasi / target * 100) if target > 0 else 0
        r_font, r_fill = get_realisasi_color(realisasi_pct_val, cumulative_target_pct)
        ws1.cell(row=r, column=17).font = r_font
        ws1.cell(row=r, column=17).fill = r_fill
            
        row_num += 1
        
    # Total row
    r = row_num
    ws1.cell(row=r, column=1, value="Kab. Kepl. Sangihe").font = Font(bold=True)
    ws1.cell(row=r, column=2, value=f"=SUM(B7:B{r-1})").font = Font(bold=True)
    ws1.cell(row=r, column=3, value=f"=SUM(C7:C{r-1})").font = Font(bold=True)
    ws1.cell(row=r, column=4, value=f"=SUM(D7:D{r-1})").font = Font(bold=True)
    ws1.cell(row=r, column=5, value=f"=D{r}/B{r}").font = Font(bold=True)
    ws1.cell(row=r, column=6, value=f"=SUM(F7:F{r-1})").font = Font(bold=True)
    ws1.cell(row=r, column=7, value=f"=F{r}/B{r}").font = Font(bold=True)
    ws1.cell(row=r, column=8, value=f"=SUM(H7:H{r-1})").font = Font(bold=True)
    ws1.cell(row=r, column=9, value=f"=H{r}/B{r}").font = Font(bold=True)
    ws1.cell(row=r, column=10, value=f"=SUM(J7:J{r-1})").font = Font(bold=True)
    ws1.cell(row=r, column=11, value=f"=J{r}/B{r}").font = Font(bold=True)
    ws1.cell(row=r, column=12, value=f"=SUM(L7:L{r-1})").font = Font(bold=True)
    ws1.cell(row=r, column=13, value=f"=L{r}/B{r}").font = Font(bold=True)
    ws1.cell(row=r, column=14, value=f"=SUM(N7:N{r-1})").font = Font(bold=True)
    ws1.cell(row=r, column=15, value=f"=N{r}/B{r}").font = Font(bold=True)
    ws1.cell(row=r, column=16, value=f"=SUM(P7:P{r-1})").font = Font(bold=True)
    ws1.cell(row=r, column=17, value=f"=P{r}/B{r}").font = Font(bold=True)
    ws1.cell(row=r, column=18, value=f"=SUM(R7:R{r-1})").font = Font(bold=True)
    ws1.cell(row=r, column=19, value=f"=R{r}/B{r}").font = Font(bold=True)
    
    for col_idx in [5, 7, 9, 11, 13, 15, 17, 19]:
        ws1.cell(row=r, column=col_idx).number_format = '0.00%'
    
    # Color the total realisasi % cell too
    total_target = sum(row["Target"] for _, row in merged_kec.iterrows())
    total_realisasi = sum(row["Realisasi_Jml"] for _, row in merged_kec.iterrows())
    total_real_pct = (total_realisasi / total_target * 100) if total_target > 0 else 0
    tr_font, tr_fill = get_realisasi_color(total_real_pct, cumulative_target_pct)
    ws1.cell(row=r, column=17).font = Font(bold=True, color=tr_font.color)
    ws1.cell(row=r, column=17).fill = tr_fill
        
    thin_border = Border(
        left=Side(style='thin', color='CCCCCC'),
        right=Side(style='thin', color='CCCCCC'),
        top=Side(style='thin', color='CCCCCC'),
        bottom=Side(style='thin', color='CCCCCC')
    )
    for row_idx in range(6, r + 1):
        for col_idx in range(1, 20):
            cell = ws1.cell(row=row_idx, column=col_idx)
            cell.border = thin_border
            if row_idx == r:
                # Keep total row orange fill, but preserve realisasi % coloring
                if col_idx != 17:
                    cell.fill = PatternFill(fill_type="solid", start_color="FCE4D6", end_color="FCE4D6")
                
    # Sheets for Leaderboards
    # Sheet 2: PPL Leaderboards
    ws2 = wb.create_sheet(title="Leaderboard PPL")
    ws2["A1"] = "LEADERBOARD PROGRES PENDATAAN LAPANGAN SE2026 MENURUT PPL"
    ws2["A1"].font = Font(size=12, bold=True)
    ws2["A3"] = f"Hari/Tanggal : {date_title}"
    ws2["A4"] = f"Pukul        : {time_title}"
    ws2["D3"] = "Target Harian:"
    ws2["D3"].font = target_label_font
    ws2["E3"] = f"{daily_target_pct}% per hari"
    ws2["E3"].font = target_value_font
    ws2["D4"] = f"Target Hari Ini (H-{elapsed_days}):"
    ws2["D4"].font = target_label_font
    ws2["E4"] = f"{cumulative_target_pct:.2f}%"
    ws2["E4"].font = target_value_font
    
    ppl_headers = ["NO", "NAMA", "KECAMATAN", "KOSEKA", "TARGET", "OPEN", "DRAFT", "SUBMIT", "REJECT", "APPROVED", "PROGRES", "REALISASI", "REALISASI (%)"]
    ws2.append([])
    ws2.append(ppl_headers)
    for col in range(1, len(ppl_headers) + 1):
        cell = ws2.cell(row=6, column=col)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")
        
    # Highest PPL
    ppl_highest = ppl_grouped.sort_values(by="Realisasi_Pct", ascending=False).reset_index(drop=True)
    row_num = 7
    for idx, row in ppl_highest.iterrows():
        r = row_num
        ws2.cell(row=r, column=1, value=idx+1)
        ws2.cell(row=r, column=2, value=row["nama_petugas"])
        ws2.cell(row=r, column=3, value=format_kec_name(row["nama_kec"]))
        ws2.cell(row=r, column=4, value=row["koseka"])
        ws2.cell(row=r, column=5, value=row["Target"])
        ws2.cell(row=r, column=6, value=row["Open"])
        ws2.cell(row=r, column=7, value=row["Draft"])
        ws2.cell(row=r, column=8, value=row["Submit"])
        ws2.cell(row=r, column=9, value=row["Reject"])
        ws2.cell(row=r, column=10, value=row["Approved"])
        ws2.cell(row=r, column=11, value=row["Progres"])
        ws2.cell(row=r, column=12, value=row["Realisasi"])
        ppl_pct_cell = ws2.cell(row=r, column=13, value=row["Realisasi_Pct"]/100)
        ppl_pct_cell.number_format = '0.00%'
        # Conditional coloring for PPL Realisasi %
        ppl_r_font, ppl_r_fill = get_realisasi_color(row["Realisasi_Pct"], cumulative_target_pct)
        ppl_pct_cell.font = ppl_r_font
        ppl_pct_cell.fill = ppl_r_fill
        row_num += 1
        
    for row_idx in range(6, row_num):
        for col_idx in range(1, 14):
            ws2.cell(row=row_idx, column=col_idx).border = thin_border
            
    # Sheet 3: PML Leaderboard
    ws3 = wb.create_sheet(title="Leaderboard PML")
    ws3["A1"] = "LEADERBOARD PROGRES PENDATAAN LAPANGAN SE2026 MENURUT PML"
    ws3["A1"].font = Font(size=12, bold=True)
    ws3["A3"] = f"Hari/Tanggal : {date_title}"
    ws3["A4"] = f"Pukul        : {time_title}"
    ws3["D3"] = "Target Harian:"
    ws3["D3"].font = target_label_font
    ws3["E3"] = f"{daily_target_pct}% per hari"
    ws3["E3"].font = target_value_font
    ws3["D4"] = f"Target Hari Ini (H-{elapsed_days}):"
    ws3["D4"].font = target_label_font
    ws3["E4"] = f"{cumulative_target_pct:.2f}%"
    ws3["E4"].font = target_value_font
    
    pml_headers = ["NO", "NAMA", "KECAMATAN", "KOSEKA", "TARGET", "OPEN", "DRAFT", "SUBMIT", "REJECT", "APPROVED", "REVOKE", "REALISASI", "REALISASI (%)"]
    ws3.append([])
    ws3.append(pml_headers)
    for col in range(1, len(pml_headers) + 1):
        cell = ws3.cell(row=6, column=col)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")
        
    pml_sorted = pml_grouped.sort_values(by="Realisasi_Pct", ascending=False).reset_index(drop=True)
    row_num = 7
    for idx, row in pml_sorted.iterrows():
        r = row_num
        ws3.cell(row=r, column=1, value=idx+1)
        ws3.cell(row=r, column=2, value=row["nama_petugas"])
        ws3.cell(row=r, column=3, value=format_kec_name(row["nama_kec"]))
        ws3.cell(row=r, column=4, value=row["koseka"])
        ws3.cell(row=r, column=5, value=row["Target"])
        ws3.cell(row=r, column=6, value=row["Open"])
        ws3.cell(row=r, column=7, value=row["Draft"])
        ws3.cell(row=r, column=8, value=row["Submit"])
        ws3.cell(row=r, column=9, value=row["Reject"])
        ws3.cell(row=r, column=10, value=row["Approved"])
        ws3.cell(row=r, column=11, value=row["Revoke"])
        ws3.cell(row=r, column=12, value=row["Realisasi"])
        pml_pct_cell = ws3.cell(row=r, column=13, value=row["Realisasi_Pct"]/100)
        pml_pct_cell.number_format = '0.00%'
        # Conditional coloring for PML Realisasi %
        pml_r_font, pml_r_fill = get_realisasi_color(row["Realisasi_Pct"], cumulative_target_pct)
        pml_pct_cell.font = pml_r_font
        pml_pct_cell.fill = pml_r_fill
        row_num += 1
        
    for row_idx in range(6, row_num):
        for col_idx in range(1, 14):
            ws3.cell(row=row_idx, column=col_idx).border = thin_border
            
    # Sheet 4: Kecamatan Leaderboard
    ws4 = wb.create_sheet(title="Leaderboard Kecamatan")
    ws4["A1"] = "LEADERBOARD PROGRES PENDATAAN LAPANGAN SE2026 MENURUT KECAMATAN"
    ws4["A1"].font = Font(size=12, bold=True)
    ws4["A3"] = f"Hari/Tanggal : {date_title}"
    ws4["A4"] = f"Pukul        : {time_title}"
    ws4["C3"] = "Target Harian:"
    ws4["C3"].font = target_label_font
    ws4["D3"] = f"{daily_target_pct}% per hari"
    ws4["D3"].font = target_value_font
    ws4["C4"] = f"Target Hari Ini (H-{elapsed_days}):"
    ws4["C4"].font = target_label_font
    ws4["D4"] = f"{cumulative_target_pct:.2f}%"
    ws4["D4"].font = target_value_font
    
    kec_headers = ["NO", "KECAMATAN", "REALISASI (%)", "RANK CHANGE"]
    ws4.append([])
    ws4.append(kec_headers)
    for col in range(1, len(kec_headers) + 1):
        cell = ws4.cell(row=6, column=col)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")
        
    row_num = 7
    for idx, row in kec_lead.iterrows():
        r = row_num
        ws4.cell(row=r, column=1, value=idx+1)
        ws4.cell(row=r, column=2, value=format_kec_name(row["nama_kec"]))
        kec_pct_cell = ws4.cell(row=r, column=3, value=row["Realisasi_Pct"]/100)
        kec_pct_cell.number_format = '0.00%'
        # Conditional coloring for Kecamatan Realisasi %
        kec_r_font, kec_r_fill = get_realisasi_color(row["Realisasi_Pct"], cumulative_target_pct)
        kec_pct_cell.font = kec_r_font
        kec_pct_cell.fill = kec_r_fill
        
        change = int(row["Rank_Change"])
        change_str = f"+{change}" if change > 0 else str(change)
        ws4.cell(row=r, column=4, value=change_str)
        
        # Color rank changes
        change_cell = ws4.cell(row=r, column=4)
        if change > 0:
            change_cell.font = Font(color="006100", bold=True) # Green
            change_cell.fill = PatternFill(fill_type="solid", start_color="C6EFCE", end_color="C6EFCE")
        elif change < 0:
            change_cell.font = Font(color="9C0006", bold=True) # Red
            change_cell.fill = PatternFill(fill_type="solid", start_color="FFC7CE", end_color="FFC7CE")
        else:
            change_cell.font = Font(color="595959") # Gray
            change_cell.fill = PatternFill(fill_type="solid", start_color="F2F2F2", end_color="F2F2F2")
            
        row_num += 1
        
    for row_idx in range(6, row_num):
        for col_idx in range(1, 5):
            ws4.cell(row=row_idx, column=col_idx).border = thin_border
            
    # Set widths for all sheets
    for ws in wb.worksheets:
        for col in ws.columns:
            max_len = max(len(str(cell.value or '')) for cell in col)
            col_letter = openpyxl.utils.get_column_letter(col[0].column)
            ws.column_dimensions[col_letter].width = max(max_len + 3, 10)
            
    # Save Report 1
    latest_path = os.path.join(public_dir, "Report_Dashboard_Latest.xlsx")
    wb.save(latest_path)
    
    # Also save specific copy for morning/evening
    specific_filename = "Report_Dashboard_Morning.xlsx" if report_type == "pagi" else "Report_Dashboard_Evening.xlsx"
    specific_path = os.path.join(public_dir, specific_filename)
    shutil.copy2(latest_path, specific_path)
    print(f"Report 1 saved to {latest_path} and {specific_path}")
    
    # Save a JSON file for the frontend to render the report page reactive tables easily
    report_json_path = os.path.join(public_dir, "report_data.json")
    
    # Convert DataFrames to dict
    kec_data_list = []
    for idx, row in merged_kec.iterrows():
        kec_data_list.append({
            "kecamatan": format_kec_name(row["nama_kec"]),
            "target": int(row["Target"]),
            "sls": int(row["SLS Code"]),
            "open": int(row["Open_Jml"]),
            "draft": int(row["Draft_Jml"]),
            "submitted": int(row["Submitted_Jml"]),
            "rejected": int(row["Rejected_Jml"]),
            "approved": int(row["Approved_Jml"]),
            "revoke": int(row["Revoke_Jml"]),
            "realisasi": int(row["Realisasi_Jml"]),
            "progres_harian": int(row["Progres_Harian"]),
            "open_pct": row["Open_Jml"] / row["Target"] if row["Target"] > 0 else 0,
            "draft_pct": row["Draft_Jml"] / row["Target"] if row["Target"] > 0 else 0,
            "submitted_pct": row["Submitted_Jml"] / row["Target"] if row["Target"] > 0 else 0,
            "rejected_pct": row["Rejected_Jml"] / row["Target"] if row["Target"] > 0 else 0,
            "approved_pct": row["Approved_Jml"] / row["Target"] if row["Target"] > 0 else 0,
            "revoke_pct": row["Revoke_Jml"] / row["Target"] if row["Target"] > 0 else 0,
            "realisasi_pct": row["Realisasi_Jml"] / row["Target"] if row["Target"] > 0 else 0,
            "progres_harian_pct": row["Progres_Harian"] / row["Target"] if row["Target"] > 0 else 0,
        })
        
    ppl_list = []
    for idx, row in ppl_grouped.iterrows():
        ppl_list.append({
            "nama": row["nama_petugas"],
            "kecamatan": format_kec_name(row["nama_kec"]),
            "koseka": row["koseka"],
            "target": int(row["Target"]),
            "open": int(row["Open"]),
            "draft": int(row["Draft"]),
            "submit": int(row["Submit"]),
            "reject": int(row["Reject"]),
            "approved": int(row["Approved"]),
            "progres": int(row["Progres"]),
            "realisasi": int(row["Realisasi"]),
            "realisasi_pct": float(row["Realisasi_Pct"])
        })
        
    pml_list = []
    for idx, row in pml_grouped.iterrows():
        pml_list.append({
            "nama": row["nama_petugas"],
            "kecamatan": format_kec_name(row["nama_kec"]),
            "koseka": row["koseka"],
            "target": int(row["Target"]),
            "open": int(row["Open"]),
            "draft": int(row["Draft"]),
            "submit": int(row["Submit"]),
            "reject": int(row["Reject"]),
            "approved": int(row["Approved"]),
            "revoke": int(row["Revoke"]),
            "realisasi": int(row["Realisasi"]),
            "realisasi_pct": float(row["Realisasi_Pct"])
        })
        
    kec_lead_list = []
    for idx, row in kec_lead.iterrows():
        kec_lead_list.append({
            "kecamatan": format_kec_name(row["nama_kec"]),
            "realisasi_pct": float(row["Realisasi_Pct"]),
            "rank": int(row["Rank"]),
            "rank_change": int(row["Rank_Change"])
        })
        
    report_type = "pagi" if hour < 13 else "sore"
    report_content = {
        "report_type": report_type,
        "date": date_title,
        "time": time_title,
        "kecamatan": kec_data_list,
        "ppl": ppl_list,
        "pml": pml_list,
        "kec_leaderboard": kec_lead_list
    }
    
    import json
    with open(report_json_path, "w", encoding="utf-8") as f:
        json.dump(report_content, f, indent=2, ensure_ascii=False)
    print("report_data.json written successfully!")
    
    # Save a specific snapshot copy
    snapshot_filename = "report_data_morning.json" if report_type == "pagi" else "report_data_evening.json"
    snapshot_path = os.path.join(public_dir, snapshot_filename)
    with open(snapshot_path, "w", encoding="utf-8") as f:
        json.dump(report_content, f, indent=2, ensure_ascii=False)
    print(f"{snapshot_filename} snapshot written successfully!")

def generate_report_2(public_dir):
    print("Generating Report 2 (Monev Excel Report)...")
    thin_border = Border(
        left=Side(style='thin', color='CCCCCC'),
        right=Side(style='thin', color='CCCCCC'),
        top=Side(style='thin', color='CCCCCC'),
        bottom=Side(style='thin', color='CCCCCC')
    )
    template_path = os.path.join("research", "fasih-dashboard-se2026", "Monev Pendataan SE2026 13 Juli 2026.xlsx")
    tabulasi_path = os.path.join("research", "fasih-dashboard-se2026", "data_mikro_sls", "tabulasi_agregat_petugas.xlsx")
    
    if not os.path.exists(template_path):
        print(f"Error: Template {template_path} not found. Skipping Report 2.")
        return
    if not os.path.exists(tabulasi_path):
        print(f"Error: {tabulasi_path} not found. Skipping Report 2.")
        return
        
    # Read tabulasi aggregates
    df_kel = pd.read_excel(tabulasi_path, sheet_name="KELUARGA")
    df_ush = pd.read_excel(tabulasi_path, sheet_name="USAHA PERUSAHAAN")
    
    # Strip spaces from column names
    df_kel.columns = [c.strip() for c in df_kel.columns]
    df_ush.columns = [c.strip() for c in df_ush.columns]
    
    # Clean email and name strings
    for df in [df_kel, df_ush]:
        df["Nama Petugas"] = df["Nama Petugas"].fillna("").astype(str).str.strip()
        df["Email Petugas"] = df["Email Petugas"].fillna("").astype(str).str.strip().str.lower()
        df["Jabatan"] = df["Jabatan"].fillna("").astype(str).str.strip()
        df["nama_kec"] = df["nama_kec"].fillna("").astype(str).str.strip()
        df["koseka"] = df["koseka"].fillna("").astype(str).str.strip()
        
    # Merge tabulasi sheets on officer details
    df_merged = pd.merge(
        df_kel, df_ush, 
        on=["Email Petugas", "Nama Petugas", "Jabatan", "nama_kec", "koseka"], 
        how="outer"
    )
    
    # Fill NaN values with appropriate zeroes
    numeric_cols = [c for c in df_merged.columns if c not in ["Email Petugas", "Nama Petugas", "Jabatan", "nama_kec", "koseka"]]
    for col in numeric_cols:
        df_merged[col] = df_merged[col].fillna(0)
        
    # Load openpyxl workbook
    wb = openpyxl.load_workbook(template_path)
    
    # Helper to clean rows from row 4 onwards
    def clean_sheet_rows(ws, start_row=4):
        max_row = ws.max_row
        if max_row >= start_row:
            ws.delete_rows(start_row, max_row - start_row + 1)
            
    # Clean and rename first sheet to current date (e.g. "18 Juli")
    current_date_str = get_current_date_str()
    for name in list(wb.sheetnames):
        if "juli" in name.lower() or name[0].isdigit():
            ws_date = wb[name]
            ws_date.title = current_date_str
            break
            
    ws_date = wb[current_date_str]
    clean_sheet_rows(ws_date, 4)
    
    # Retrieve row 4 style from the original sheet before deletion
    # We load template path again just to get the original row styles
    wb_orig = openpyxl.load_workbook(template_path)
    ws_orig = wb_orig["13 Juli"]
    
    row_style_cells = [ws_orig.cell(row=4, column=c) for c in range(1, 42)]
    pml_row_style_cells = [ws_orig.cell(row=152, column=c) for c in range(1, 42)] # row 152 is PML row style
    
    # Populate sheet "18 Juli" (sorted alphabetically by officer name)
    df_all_sorted = df_merged.sort_values(by="Nama Petugas").reset_index(drop=True)
    
    row_num = 4
    for idx, row in df_all_sorted.iterrows():
        r = row_num
        is_pml = row["Jabatan"] == "PML"
        style_source = pml_row_style_cells if is_pml else row_style_cells
        
        # Target Prelist Eligible
        f_val = int(row.get("Prelist Awal", 0))
        g_val = int(row.get("Jumlah Prelist Usaha", 0))
        
        # Tidak Didata
        meninggal = int(row.get("Meninggal", 0))
        tidak_eligible = int(row.get("Tidak Eligible", 0))
        tidak_dapat_ditemui = int(row.get("Tidak Dapat Ditemui Sampai Akhir Pendataan", 0))
        tidak_ditemukan_kel = int(row.get("Tidak Ditemukan", 0))
        
        i_val = meninggal + tidak_eligible + tidak_dapat_ditemui + tidak_ditemukan_kel
        
        tutup = int(row.get("JUMLAH USAHA BKU MENURUT STATUS KEBERADAAN USAHA - Tutup", 0))
        ganda = int(row.get("JUMLAH USAHA BKU MENURUT STATUS KEBERADAAN USAHA - Ganda", 0))
        tidak_ditemukan_ush = int(row.get("JUMLAH USAHA BKU MENURUT STATUS KEBERADAAN USAHA - Tidak Ditemukan", 0))
        
        j_val = tutup + ganda + tidak_ditemukan_ush
        
        # Status Muatan Pendataan
        o_val = i_val
        p_val = int(row.get("Keluarga Baru", 0))
        r_val = j_val
        s_val = int(row.get("JUMLAH USAHA BKU MENURUT STATUS KEBERADAAN USAHA - Baru", 0))
        
        # Realisasi
        ditemukan_kel = int(row.get("Ditemukan", 0))
        x_val = ditemukan_kel + i_val
        
        usaha_bku = int(row.get("Jumlah Usaha BKU", 0))
        y_val = usaha_bku + j_val
        
        aa_val = ditemukan_kel
        ab_val = usaha_bku
        
        # Write values
        ws_date.cell(row=r, column=1, value=idx+1)
        ws_date.cell(row=r, column=2, value=row["Nama Petugas"])
        ws_date.cell(row=r, column=3, value=row["Email Petugas"])
        ws_date.cell(row=r, column=4, value=row["Jabatan"])
        ws_date.cell(row=r, column=5, value=row["nama_kec"])
        
        ws_date.cell(row=r, column=6, value=f_val)
        ws_date.cell(row=r, column=7, value=g_val)
        ws_date.cell(row=r, column=8, value=f"=F{r}+G{r}")
        
        ws_date.cell(row=r, column=9, value=i_val)
        ws_date.cell(row=r, column=10, value=j_val)
        ws_date.cell(row=r, column=11, value=f"=I{r}+J{r}")
        
        ws_date.cell(row=r, column=12, value=f"=F{r}-I{r}")
        ws_date.cell(row=r, column=13, value=f"=G{r}-J{r}")
        ws_date.cell(row=r, column=14, value=f"=L{r}+M{r}")
        
        ws_date.cell(row=r, column=15, value=o_val)
        ws_date.cell(row=r, column=16, value=p_val)
        ws_date.cell(row=r, column=17, value=f"=P{r}-O{r}")
        
        ws_date.cell(row=r, column=18, value=r_val)
        ws_date.cell(row=r, column=19, value=s_val)
        ws_date.cell(row=r, column=20, value=f"=S{r}-R{r}")
        
        ws_date.cell(row=r, column=21, value=f"=O{r}+R{r}")
        ws_date.cell(row=r, column=22, value=f"=P{r}+S{r}")
        ws_date.cell(row=r, column=23, value=f"=V{r}-U{r}")
        
        ws_date.cell(row=r, column=24, value=x_val)
        ws_date.cell(row=r, column=25, value=y_val)
        ws_date.cell(row=r, column=26, value=f"=X{r}+Y{r}")
        
        ws_date.cell(row=r, column=27, value=aa_val)
        ws_date.cell(row=r, column=28, value=ab_val)
        ws_date.cell(row=r, column=29, value=f"=AA{r}+AB{r}")
        
        ws_date.cell(row=r, column=30, value=f"=X{r}/F{r}")
        ws_date.cell(row=r, column=31, value=f"=Y{r}/G{r}")
        ws_date.cell(row=r, column=32, value=f"=Z{r}/H{r}")
        
        ws_date.cell(row=r, column=33, value=f"=AA{r}/F{r}")
        ws_date.cell(row=r, column=34, value=f"=AB{r}/G{r}")
        ws_date.cell(row=r, column=35, value=f"=AC{r}/H{r}")
        
        ws_date.cell(row=r, column=36, value=f"=X{r}/L{r}")
        ws_date.cell(row=r, column=37, value=f"=Y{r}/M{r}")
        ws_date.cell(row=r, column=38, value=f"=Z{r}/N{r}")
        
        ws_date.cell(row=r, column=39, value=f"=AA{r}/L{r}")
        ws_date.cell(row=r, column=40, value=f"=AB{r}/M{r}")
        ws_date.cell(row=r, column=41, value=f"=AC{r}/N{r}")
        
        # Apply styles and alignments
        for c in range(1, 42):
            cell = ws_date.cell(row=r, column=c)
            copy_cell_style(style_source[c-1], cell)
            
        row_num += 1
        
    print(f"Sheet {current_date_str} written with {row_num-4} rows.")
    
    # We will write similar loops to populate other sheets
    # Helper to calculate and sort for PPL/PML sheets
    # We can pre-calculate the progress values in Python to sort them properly
    ppl_rows = []
    pml_rows = []
    for idx, row in df_merged.iterrows():
        # Target
        f_val = int(row.get("Prelist Awal", 0))
        g_val = int(row.get("Jumlah Prelist Usaha", 0))
        h_val = f_val + g_val
        
        # Tidak Didata
        meninggal = int(row.get("Meninggal", 0))
        tidak_eligible = int(row.get("Tidak Eligible", 0))
        tidak_dapat_ditemui = int(row.get("Tidak Dapat Ditemui Sampai Akhir Pendataan", 0))
        tidak_ditemukan_kel = int(row.get("Tidak Ditemukan", 0))
        i_val = meninggal + tidak_eligible + tidak_dapat_ditemui + tidak_ditemukan_kel
        
        tutup = int(row.get("JUMLAH USAHA BKU MENURUT STATUS KEBERADAAN USAHA - Tutup", 0))
        ganda = int(row.get("JUMLAH USAHA BKU MENURUT STATUS KEBERADAAN USAHA - Ganda", 0))
        tidak_ditemukan_ush = int(row.get("JUMLAH USAHA BKU MENURUT STATUS KEBERADAAN USAHA - Tidak Ditemukan", 0))
        j_val = tutup + ganda + tidak_ditemukan_ush
        k_val = i_val + j_val
        
        # Realisasi
        ditemukan_kel = int(row.get("Ditemukan", 0))
        x_val = ditemukan_kel + i_val
        usaha_bku = int(row.get("Jumlah Usaha BKU", 0))
        y_val = usaha_bku + j_val
        z_val = x_val + y_val
        
        aa_val = ditemukan_kel
        ab_val = usaha_bku
        ac_val = aa_val + ab_val
        
        progress_pct = z_val / h_val if h_val > 0 else 0
        ush_b = int(row.get("JUMLAH USAHA BKU MENURUT STATUS KEBERADAAN USAHA - Baru", 0))
        ush_sel = ush_b - j_val
        
        obj = {
            'row': row,
            'f_val': f_val, 'g_val': g_val, 'h_val': h_val,
            'i_val': i_val, 'j_val': j_val, 'k_val': k_val,
            'l_val': f_val - i_val, 'm_val': g_val - j_val, 'n_val': (f_val - i_val) + (g_val - j_val),
            'o_val': i_val, 'p_val': int(row.get("Keluarga Baru", 0)),
            'r_val': j_val, 's_val': ush_b,
            'x_val': x_val, 'y_val': y_val, 'z_val': z_val,
            'aa_val': aa_val, 'ab_val': ab_val, 'ac_val': ac_val,
            'progress_pct': progress_pct,
            'ush_sel': ush_sel
        }
        if row["Jabatan"] == "PML":
            pml_rows.append(obj)
        else:
            ppl_rows.append(obj)
            
    # Sort sheets
    # 1. Progres PPL (sorted by progress_pct descending)
    ppl_sorted_progress = sorted(ppl_rows, key=lambda x: (-x['progress_pct'], x['row']['Nama Petugas']))
    ws_ppl = wb["Progres PPL"]
    clean_sheet_rows(ws_ppl, 4)
    row_num = 4
    for idx, obj in enumerate(ppl_sorted_progress):
        r = row_num
        ws_ppl.cell(row=r, column=1, value=idx+1)
        ws_ppl.cell(row=r, column=2, value=obj['row']["Nama Petugas"])
        ws_ppl.cell(row=r, column=3, value=obj['row']["Email Petugas"])
        ws_ppl.cell(row=r, column=4, value=obj['row']["Jabatan"])
        ws_ppl.cell(row=r, column=5, value=obj['row']["nama_kec"])
        
        ws_ppl.cell(row=r, column=6, value=obj['f_val'])
        ws_ppl.cell(row=r, column=7, value=obj['g_val'])
        ws_ppl.cell(row=r, column=8, value=f"=F{r}+G{r}")
        ws_ppl.cell(row=r, column=9, value=obj['i_val'])
        ws_ppl.cell(row=r, column=10, value=obj['j_val'])
        ws_ppl.cell(row=r, column=11, value=f"=I{r}+J{r}")
        ws_ppl.cell(row=r, column=12, value=f"=F{r}-I{r}")
        ws_ppl.cell(row=r, column=13, value=f"=G{r}-J{r}")
        ws_ppl.cell(row=r, column=14, value=f"=L{r}+M{r}")
        ws_ppl.cell(row=r, column=15, value=obj['o_val'])
        ws_ppl.cell(row=r, column=16, value=obj['p_val'])
        ws_ppl.cell(row=r, column=17, value=f"=P{r}-O{r}")
        ws_ppl.cell(row=r, column=18, value=obj['r_val'])
        ws_ppl.cell(row=r, column=19, value=obj['s_val'])
        ws_ppl.cell(row=r, column=20, value=f"=S{r}-R{r}")
        ws_ppl.cell(row=r, column=21, value=f"=O{r}+R{r}")
        ws_ppl.cell(row=r, column=22, value=f"=P{r}+S{r}")
        ws_ppl.cell(row=r, column=23, value=f"=V{r}-U{r}")
        ws_ppl.cell(row=r, column=24, value=obj['x_val'])
        ws_ppl.cell(row=r, column=25, value=obj['y_val'])
        ws_ppl.cell(row=r, column=26, value=f"=X{r}+Y{r}")
        ws_ppl.cell(row=r, column=27, value=obj['aa_val'])
        ws_ppl.cell(row=r, column=28, value=obj['ab_val'])
        ws_ppl.cell(row=r, column=29, value=f"=AA{r}+AB{r}")
        
        ws_ppl.cell(row=r, column=30, value=f"=X{r}/F{r}")
        ws_ppl.cell(row=r, column=31, value=f"=Y{r}/G{r}")
        ws_ppl.cell(row=r, column=32, value=f"=Z{r}/H{r}")
        ws_ppl.cell(row=r, column=33, value=f"=AA{r}/F{r}")
        ws_ppl.cell(row=r, column=34, value=f"=AB{r}/G{r}")
        ws_ppl.cell(row=r, column=35, value=f"=AC{r}/H{r}")
        ws_ppl.cell(row=r, column=36, value=f"=X{r}/L{r}")
        ws_ppl.cell(row=r, column=37, value=f"=Y{r}/M{r}")
        ws_ppl.cell(row=r, column=38, value=f"=Z{r}/N{r}")
        ws_ppl.cell(row=r, column=39, value=f"=AA{r}/L{r}")
        ws_ppl.cell(row=r, column=40, value=f"=AB{r}/M{r}")
        ws_ppl.cell(row=r, column=41, value=f"=AC{r}/N{r}")
        
        for c in range(1, 42):
            cell = ws_ppl.cell(row=r, column=c)
            copy_cell_style(row_style_cells[c-1], cell)
        row_num += 1
        
    print(f"Sheet Progres PPL written with {row_num-4} rows.")
    
    # 2. Progres PML (sorted by progress_pct descending)
    pml_sorted = sorted(pml_rows, key=lambda x: (-x['progress_pct'], x['row']['Nama Petugas']))
    ws_pml = wb["Progres PML"]
    clean_sheet_rows(ws_pml, 4)
    row_num = 4
    for idx, obj in enumerate(pml_sorted):
        r = row_num
        ws_pml.cell(row=r, column=1, value=idx+1)
        ws_pml.cell(row=r, column=2, value=obj['row']["Nama Petugas"])
        ws_pml.cell(row=r, column=3, value=obj['row']["Email Petugas"])
        ws_pml.cell(row=r, column=4, value=obj['row']["Jabatan"])
        ws_pml.cell(row=r, column=5, value=obj['row']["nama_kec"])
        
        ws_pml.cell(row=r, column=6, value=obj['f_val'])
        ws_pml.cell(row=r, column=7, value=obj['g_val'])
        ws_pml.cell(row=r, column=8, value=f"=F{r}+G{r}")
        ws_pml.cell(row=r, column=9, value=obj['i_val'])
        ws_pml.cell(row=r, column=10, value=obj['j_val'])
        ws_pml.cell(row=r, column=11, value=f"=I{r}+J{r}")
        ws_pml.cell(row=r, column=12, value=f"=F{r}-I{r}")
        ws_pml.cell(row=r, column=13, value=f"=G{r}-J{r}")
        ws_pml.cell(row=r, column=14, value=f"=L{r}+M{r}")
        ws_pml.cell(row=r, column=15, value=obj['o_val'])
        ws_pml.cell(row=r, column=16, value=obj['p_val'])
        ws_pml.cell(row=r, column=17, value=f"=P{r}-O{r}")
        ws_pml.cell(row=r, column=18, value=obj['r_val'])
        ws_pml.cell(row=r, column=19, value=obj['s_val'])
        ws_pml.cell(row=r, column=20, value=f"=S{r}-R{r}")
        ws_pml.cell(row=r, column=21, value=f"=O{r}+R{r}")
        ws_pml.cell(row=r, column=22, value=f"=P{r}+S{r}")
        ws_pml.cell(row=r, column=23, value=f"=V{r}-U{r}")
        ws_pml.cell(row=r, column=24, value=obj['x_val'])
        ws_pml.cell(row=r, column=25, value=obj['y_val'])
        ws_pml.cell(row=r, column=26, value=f"=X{r}+Y{r}")
        ws_pml.cell(row=r, column=27, value=obj['aa_val'])
        ws_pml.cell(row=r, column=28, value=obj['ab_val'])
        ws_pml.cell(row=r, column=29, value=f"=AA{r}+AB{r}")
        
        ws_pml.cell(row=r, column=30, value=f"=X{r}/F{r}")
        ws_pml.cell(row=r, column=31, value=f"=Y{r}/G{r}")
        ws_pml.cell(row=r, column=32, value=f"=Z{r}/H{r}")
        ws_pml.cell(row=r, column=33, value=f"=AA{r}/F{r}")
        ws_pml.cell(row=r, column=34, value=f"=AB{r}/G{r}")
        ws_pml.cell(row=r, column=35, value=f"=AC{r}/H{r}")
        ws_pml.cell(row=r, column=36, value=f"=X{r}/L{r}")
        ws_pml.cell(row=r, column=37, value=f"=Y{r}/M{r}")
        ws_pml.cell(row=r, column=38, value=f"=Z{r}/N{r}")
        ws_pml.cell(row=r, column=39, value=f"=AA{r}/L{r}")
        ws_pml.cell(row=r, column=40, value=f"=AB{r}/M{r}")
        ws_pml.cell(row=r, column=41, value=f"=AC{r}/N{r}")
        
        for c in range(1, 42):
            cell = ws_pml.cell(row=r, column=c)
            copy_cell_style(pml_row_style_cells[c-1], cell)
        row_num += 1
        
    print(f"Sheet Progres PML written with {row_num-4} rows.")
    
    # 3. Status Muatan (sorted by ush_sel ascending)
    ppl_sorted_selisih = sorted(ppl_rows, key=lambda x: (x['ush_sel'], x['row']['Nama Petugas']))
    ws_muatan = wb["Status Muatan"]
    clean_sheet_rows(ws_muatan, 4)
    row_num = 4
    for idx, obj in enumerate(ppl_sorted_selisih):
        r = row_num
        ws_muatan.cell(row=r, column=1, value=idx+1)
        ws_muatan.cell(row=r, column=2, value=obj['row']["Nama Petugas"])
        ws_muatan.cell(row=r, column=3, value=obj['row']["Email Petugas"])
        ws_muatan.cell(row=r, column=4, value=obj['row']["Jabatan"])
        ws_muatan.cell(row=r, column=5, value=obj['row']["nama_kec"])
        
        ws_muatan.cell(row=r, column=6, value=obj['f_val'])
        ws_muatan.cell(row=r, column=7, value=obj['g_val'])
        ws_muatan.cell(row=r, column=8, value=f"=F{r}+G{r}")
        ws_muatan.cell(row=r, column=9, value=obj['i_val'])
        ws_muatan.cell(row=r, column=10, value=obj['j_val'])
        ws_muatan.cell(row=r, column=11, value=f"=I{r}+J{r}")
        ws_muatan.cell(row=r, column=12, value=f"=F{r}-I{r}")
        ws_muatan.cell(row=r, column=13, value=f"=G{r}-J{r}")
        ws_muatan.cell(row=r, column=14, value=f"=L{r}+M{r}")
        ws_muatan.cell(row=r, column=15, value=obj['o_val'])
        ws_muatan.cell(row=r, column=16, value=obj['p_val'])
        ws_muatan.cell(row=r, column=17, value=f"=P{r}-O{r}")
        ws_muatan.cell(row=r, column=18, value=obj['r_val'])
        ws_muatan.cell(row=r, column=19, value=obj['s_val'])
        ws_muatan.cell(row=r, column=20, value=f"=S{r}-R{r}")
        ws_muatan.cell(row=r, column=21, value=f"=O{r}+R{r}")
        ws_muatan.cell(row=r, column=22, value=f"=P{r}+S{r}")
        ws_muatan.cell(row=r, column=23, value=f"=V{r}-U{r}")
        ws_muatan.cell(row=r, column=24, value=obj['x_val'])
        ws_muatan.cell(row=r, column=25, value=obj['y_val'])
        ws_muatan.cell(row=r, column=26, value=f"=X{r}+Y{r}")
        ws_muatan.cell(row=r, column=27, value=obj['aa_val'])
        ws_muatan.cell(row=r, column=28, value=obj['ab_val'])
        ws_muatan.cell(row=r, column=29, value=f"=AA{r}+AB{r}")
        
        ws_muatan.cell(row=r, column=30, value=f"=X{r}/F{r}")
        ws_muatan.cell(row=r, column=31, value=f"=Y{r}/G{r}")
        ws_muatan.cell(row=r, column=32, value=f"=Z{r}/H{r}")
        ws_muatan.cell(row=r, column=33, value=f"=AA{r}/F{r}")
        ws_muatan.cell(row=r, column=34, value=f"=AB{r}/G{r}")
        ws_muatan.cell(row=r, column=35, value=f"=AC{r}/H{r}")
        ws_muatan.cell(row=r, column=36, value=f"=X{r}/L{r}")
        ws_muatan.cell(row=r, column=37, value=f"=Y{r}/M{r}")
        ws_muatan.cell(row=r, column=38, value=f"=Z{r}/N{r}")
        ws_muatan.cell(row=r, column=39, value=f"=AA{r}/L{r}")
        ws_muatan.cell(row=r, column=40, value=f"=AB{r}/M{r}")
        ws_muatan.cell(row=r, column=41, value=f"=AC{r}/N{r}")
        
        for c in range(1, 42):
            cell = ws_muatan.cell(row=r, column=c)
            copy_cell_style(row_style_cells[c-1], cell)
        row_num += 1
        
    print(f"Sheet Status Muatan written with {row_num-4} rows.")
    
    # 4. Kecamatan (grouped by Kecamatan, preserving original names and order)
    ws_kec = wb["Kecamatan"]
    
    kec_mapping = {
        "Manganitu Selatan": ["(040) MANGANITU SELATAN"],
        "Tatoareng": ["(041) TATOARENG"],
        "Tamako": ["(050) TAMAKO"],
        "Tamako dan Manganitu Selatan": ["(050) TAMAKO & (080) MANGANITU"],
        "Tabukan Selatan": ["(060) TABUKAN SELATAN"],
        "Tabukan Selatan dan Tabukan Selatan Tengah": ["(060) TABUKAN SELATAN & (070) TABUKAN TENGAH"],
        "Tabukan Selatan Tengah": ["(061) TABUKAN SELATAN TENGAH"],
        "Tabukan Selatan Tengah dan Tabukan Selatan Tenggara": ["(061) TABUKAN SELATAN TENGAH & (062) TABUKAN SELATAN TENGGARA"],
        "Tabukan Tengah": ["(070) TABUKAN TENGAH"],
        "Manganitu": ["(080) MANGANITU"],
        "Tahuna": ["(090) TAHUNA"],
        "Tahuna Timur": ["(091) TAHUNA TIMUR"],
        "Tahuna Barat": ["(092) TAHUNA BARAT"],
        "Tabukan Utara": ["(100) TABUKAN UTARA"],
        "Tabukan Utara dan Nusa Tabukan": ["(100) TABUKAN UTARA & (101) NUSA TABUKAN"],
        "Tabukan Utara dan Kepl. Marore": ["(100) TABUKAN UTARA & (102) KEPULAUAN MARORE"],
        "Nusa Tabukan": ["(101) NUSA TABUKAN"],
        "Kepulauan Marore": ["(102) KEPULAUAN MARORE"],
        "Kendahe": ["(110) KENDAHE"]
    }
    
    # We overwrite rows 5 to 23 of Kecamatan sheet.
    # Note: Column A is No, Column B is KECAMATAN name
    for r in range(5, 24):
        kec_name = ws_kec.cell(row=r, column=2).value
        mapped_kec_codes = kec_mapping.get(kec_name, [])
        
        # Filter df_merged for these codes
        df_sub = df_merged[df_merged["nama_kec"].isin(mapped_kec_codes)]
        
        # Target
        f_val = sum(df_sub["Prelist Awal"].fillna(0).astype(int))
        g_val = sum(df_sub["Jumlah Prelist Usaha"].fillna(0).astype(int))
        
        # Tidak Didata
        meninggal = sum(df_sub["Meninggal"].fillna(0).astype(int))
        tidak_eligible = sum(df_sub["Tidak Eligible"].fillna(0).astype(int))
        tidak_dapat_ditemui = sum(df_sub["Tidak Dapat Ditemui Sampai Akhir Pendataan"].fillna(0).astype(int))
        tidak_ditemukan_kel = sum(df_sub["Tidak Ditemukan"].fillna(0).astype(int))
        i_val = meninggal + tidak_eligible + tidak_dapat_ditemui + tidak_ditemukan_kel
        
        tutup = sum(df_sub["JUMLAH USAHA BKU MENURUT STATUS KEBERADAAN USAHA - Tutup"].fillna(0).astype(int))
        ganda = sum(df_sub["JUMLAH USAHA BKU MENURUT STATUS KEBERADAAN USAHA - Ganda"].fillna(0).astype(int))
        tidak_ditemukan_ush = sum(df_sub["JUMLAH USAHA BKU MENURUT STATUS KEBERADAAN USAHA - Tidak Ditemukan"].fillna(0).astype(int))
        j_val = tutup + ganda + tidak_ditemukan_ush
        
        # Status Muatan Pendataan
        o_val = i_val
        p_val = sum(df_sub["Keluarga Baru"].fillna(0).astype(int))
        r_val = j_val
        s_val = sum(df_sub["JUMLAH USAHA BKU MENURUT STATUS KEBERADAAN USAHA - Baru"].fillna(0).astype(int))
        
        # Realisasi
        ditemukan_kel = sum(df_sub["Ditemukan"].fillna(0).astype(int))
        x_val = ditemukan_kel + i_val
        usaha_bku = sum(df_sub["Jumlah Usaha BKU"].fillna(0).astype(int))
        y_val = usaha_bku + j_val
        
        aa_val = ditemukan_kel
        ab_val = usaha_bku
        
        # Write values
        ws_kec.cell(row=r, column=3, value=f_val)
        ws_kec.cell(row=r, column=4, value=g_val)
        ws_kec.cell(row=r, column=5, value=f"=C{r}+D{r}")
        
        ws_kec.cell(row=r, column=6, value=i_val)
        ws_kec.cell(row=r, column=7, value=j_val)
        ws_kec.cell(row=r, column=8, value=f"=F{r}+G{r}")
        
        ws_kec.cell(row=r, column=9, value=f"=C{r}-F{r}")
        ws_kec.cell(row=r, column=10, value=f"=D{r}-G{r}")
        ws_kec.cell(row=r, column=11, value=f"=I{r}+J{r}")
        
        ws_kec.cell(row=r, column=12, value=o_val)
        ws_kec.cell(row=r, column=13, value=p_val)
        ws_kec.cell(row=r, column=14, value=f"=M{r}-L{r}")
        
        ws_kec.cell(row=r, column=15, value=r_val)
        ws_kec.cell(row=r, column=16, value=s_val)
        ws_kec.cell(row=r, column=17, value=f"=P{r}-O{r}")
        
        ws_kec.cell(row=r, column=18, value=f"=L{r}+O{r}")
        ws_kec.cell(row=r, column=19, value=f"=M{r}+P{r}")
        ws_kec.cell(row=r, column=20, value=f"=S{r}-R{r}")
        
        ws_kec.cell(row=r, column=21, value=x_val)
        ws_kec.cell(row=r, column=22, value=y_val)
        ws_kec.cell(row=r, column=23, value=f"=U{r}+V{r}")
        
        ws_kec.cell(row=r, column=24, value=aa_val)
        ws_kec.cell(row=r, column=25, value=ab_val)
        ws_kec.cell(row=r, column=26, value=f"=X{r}+Y{r}")
        
        ws_kec.cell(row=r, column=27, value=f"=U{r}/C{r}")
        ws_kec.cell(row=r, column=28, value=f"=V{r}/D{r}")
        ws_kec.cell(row=r, column=29, value=f"=W{r}/E{r}")
        
        ws_kec.cell(row=r, column=30, value=f"=X{r}/C{r}")
        ws_kec.cell(row=r, column=31, value=f"=Y{r}/D{r}")
        ws_kec.cell(row=r, column=32, value=f"=Z{r}/E{r}")
        
        ws_kec.cell(row=r, column=33, value=f"=U{r}/I{r}")
        ws_kec.cell(row=r, column=34, value=f"=V{r}/J{r}")
        ws_kec.cell(row=r, column=35, value=f"=W{r}/K{r}")
        
        ws_kec.cell(row=r, column=36, value=f"=X{r}/I{r}")
        ws_kec.cell(row=r, column=37, value=f"=Y{r}/J{r}")
        ws_kec.cell(row=r, column=38, value=f"=Z{r}/K{r}")
        
    print("Sheet Kecamatan updated successfully!")
    
    # 5. Rekap Kecamatan (new sheet, only Kecamatan name and progress percentage)
    ws_rekap = wb.create_sheet(title="Rekap Kecamatan")
    ws_rekap["A1"] = "REKAPITULASI PROGRES PENDATAAN LAPANGAN SE2026 PER KECAMATAN"
    ws_rekap["A1"].font = Font(name="Calibri", size=12, bold=True)
    
    rekap_headers = ["NO", "KECAMATAN", "PROGRES (%)"]
    ws_rekap.append([])
    ws_rekap.append(rekap_headers)
    for col in range(1, len(rekap_headers) + 1):
        cell = ws_rekap.cell(row=3, column=col)
        cell.fill = PatternFill(fill_type="solid", start_color="1F497D", end_color="1F497D") # Navy blue
        cell.font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
        cell.alignment = Alignment(horizontal="center")
        
    # We populate the 15 single subdistricts
    rekap_list = []
    rekap_kec_names = [
        "Manganitu Selatan", "Tatoareng", "Tamako", "Tabukan Selatan", 
        "Tabukan Selatan Tengah", "Tabukan Tengah", "Manganitu", "Tahuna", 
        "Tahuna Timur", "Tahuna Barat", "Tabukan Utara", "Nusa Tabukan", 
        "Kepulauan Marore", "Kendahe"
    ]
    
    row_num = 4
    for idx, kec_name in enumerate(rekap_kec_names):
        r = row_num
        mapped_codes = kec_mapping.get(kec_name, [])
        df_sub = df_merged[df_merged["nama_kec"].isin(mapped_codes)]
        
        # Calculate progress
        target_kel = sum(df_sub["Prelist Awal"].fillna(0).astype(int))
        target_ush = sum(df_sub["Jumlah Prelist Usaha"].fillna(0).astype(int))
        total_target = target_kel + target_ush
        
        # Realisasi
        meninggal = sum(df_sub["Meninggal"].fillna(0).astype(int))
        tidak_eligible = sum(df_sub["Tidak Eligible"].fillna(0).astype(int))
        tidak_dapat_ditemui = sum(df_sub["Tidak Dapat Ditemui Sampai Akhir Pendataan"].fillna(0).astype(int))
        tidak_ditemukan_kel = sum(df_sub["Tidak Ditemukan"].fillna(0).astype(int))
        i_val = meninggal + tidak_eligible + tidak_dapat_ditemui + tidak_ditemukan_kel
        ditemukan_kel = sum(df_sub["Ditemukan"].fillna(0).astype(int))
        x_val = ditemukan_kel + i_val
        
        tutup = sum(df_sub["JUMLAH USAHA BKU MENURUT STATUS KEBERADAAN USAHA - Tutup"].fillna(0).astype(int))
        ganda = sum(df_sub["JUMLAH USAHA BKU MENURUT STATUS KEBERADAAN USAHA - Ganda"].fillna(0).astype(int))
        tidak_ditemukan_ush = sum(df_sub["JUMLAH USAHA BKU MENURUT STATUS KEBERADAAN USAHA - Tidak Ditemukan"].fillna(0).astype(int))
        j_val = tutup + ganda + tidak_ditemukan_ush
        usaha_bku = sum(df_sub["Jumlah Usaha BKU"].fillna(0).astype(int))
        y_val = usaha_bku + j_val
        
        realisasi_total = x_val + y_val
        progress_pct = realisasi_total / total_target if total_target > 0 else 0
        
        rekap_list.append({
            "kecamatan": kec_name,
            "progress_pct": float(progress_pct)
        })
        
        ws_rekap.cell(row=r, column=1, value=idx+1)
        ws_rekap.cell(row=r, column=2, value=kec_name)
        ws_rekap.cell(row=r, column=3, value=progress_pct).number_format = '0.00%'
        row_num += 1
        
    # Styling for Rekap sheet
    for row_idx in range(3, row_num):
        for col_idx in range(1, 4):
            ws_rekap.cell(row=row_idx, column=col_idx).border = thin_border
            ws_rekap.cell(row=row_idx, column=col_idx).font = Font(name="Calibri", size=11)
            if row_idx % 2 == 1 and row_idx > 3:
                ws_rekap.cell(row=row_idx, column=col_idx).fill = PatternFill(fill_type="solid", start_color="F2F2F2", end_color="F2F2F2")
                
    ws_rekap.column_dimensions["A"].width = 6
    ws_rekap.column_dimensions["B"].width = 25
    ws_rekap.column_dimensions["C"].width = 18
    
    print("Sheet Rekap Kecamatan created successfully!")
    
    # Save Report 2
    date_formatted = START_TIME.strftime("%d %B %Y")
    # map English month to Indonesian
    months_en_id = {
        "January": "Januari", "February": "Februari", "March": "Maret", "April": "April",
        "May": "Mei", "June": "Juni", "July": "Juli", "August": "Agustus",
        "September": "September", "October": "Oktober", "November": "November", "December": "Desember"
    }
    for en, ind in months_en_id.items():
        date_formatted = date_formatted.replace(en, ind)
        
    dest_path_research = os.path.join("research", "fasih-dashboard-se2026", f"Monev Pendataan SE2026 {date_formatted}.xlsx")
    dest_path_public = os.path.join(public_dir, "Monev_Pendataan_SE2026_Latest.xlsx")
    wb.save(dest_path_public)
    
    # Also save specific copy for morning/evening
    hour = START_TIME.hour
    report_type = "pagi" if hour < 13 else "sore"
    specific_filename = "Monev_Pendataan_SE2026_Morning.xlsx" if report_type == "pagi" else "Monev_Pendataan_SE2026_Evening.xlsx"
    specific_path = os.path.join(public_dir, specific_filename)
    shutil.copy2(dest_path_public, specific_path)
    
    shutil.copy2(dest_path_public, dest_path_research)
    print(f"Report 2 saved to {dest_path_public}, {specific_path} and {dest_path_research}")
    
    # Save Rekap as JSON for the frontend
    import json
    rekap_json_path = os.path.join(public_dir, "monev_rekap.json")
    with open(rekap_json_path, "w", encoding="utf-8") as f:
        json.dump(rekap_list, f, indent=2, ensure_ascii=False)
    print("monev_rekap.json saved successfully!")
    
    # Update report_data.json with status_muatan
    report_json_path = os.path.join(public_dir, "report_data.json")
    if os.path.exists(report_json_path):
        try:
            with open(report_json_path, "r", encoding="utf-8") as f:
                report_data = json.load(f)
            
            # Combine PPL and PML rows
            all_muatan_rows = ppl_rows + pml_rows
            all_sorted_selisih = sorted(all_muatan_rows, key=lambda x: (x['ush_sel'], x['row']['Nama Petugas']))
            
            status_muatan_list = []
            for idx, obj in enumerate(all_sorted_selisih):
                row = obj['row']
                status_muatan_list.append({
                    "rank": idx + 1,
                    "nama": row["Nama Petugas"],
                    "jabatan": "PML" if row["Jabatan"] == "PML" else "PCL",
                    "kecamatan": format_kec_name(row["nama_kec"]),
                    "kel_td": int(obj['o_val']),
                    "kel_b": int(obj['p_val']),
                    "kel_sel": int(obj['p_val'] - obj['o_val']),
                    "ush_td": int(obj['r_val']),
                    "ush_b": int(obj['s_val']),
                    "ush_sel": int(obj['ush_sel']),
                    "tot_td": int(obj['o_val'] + obj['r_val']),
                    "tot_b": int(obj['p_val'] + obj['s_val']),
                    "tot_sel": int((obj['p_val'] + obj['s_val']) - (obj['o_val'] + obj['r_val']))
                })
            
            report_data["status_muatan"] = status_muatan_list
            report_type = report_data.get("report_type", "pagi")
            
            with open(report_json_path, "w", encoding="utf-8") as f:
                json.dump(report_data, f, indent=2, ensure_ascii=False)
            print("report_data.json updated with status_muatan successfully!")
            
            # Also update snapshot
            snapshot_filename = "report_data_morning.json" if report_type == "pagi" else "report_data_evening.json"
            snapshot_path = os.path.join(public_dir, snapshot_filename)
            with open(snapshot_path, "w", encoding="utf-8") as f:
                json.dump(report_data, f, indent=2, ensure_ascii=False)
            print(f"{snapshot_filename} updated with status_muatan successfully!")
        except Exception as e:
            print(f"Error updating report_data.json: {e}")


def main():
    public_dir = os.path.join("dashboard", "public")
    os.makedirs(public_dir, exist_ok=True)
    
    generate_report_1(public_dir)
    generate_report_2(public_dir)
    print("All reports generated successfully!")

if __name__ == "__main__":
    main()
