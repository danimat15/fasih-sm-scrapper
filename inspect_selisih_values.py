import openpyxl
import os

def main():
    excel_path = os.path.join("research", "fasih-dashboard-se2026", "Monev Pendataan SE2026 13 Juli 2026.xlsx")
    if not os.path.exists(excel_path):
        print("Monev file not found!")
        return

    wb = openpyxl.load_workbook(excel_path, data_only=True)
    ws = wb["Status Muatan"]
    
    print("Monev Status Muatan Sheet rows:")
    for r in range(2, 40): # start at 2 since data starts there
        no = ws.cell(row=r, column=1).value
        name = ws.cell(row=r, column=2).value
        
        # O: Kel Tidak Didata (15), P: Kel Baru (16), Q: Kel Selisih (17)
        # R: Usaha Tidak Didata (18), S: Usaha Baru (19), T: Usaha Selisih (20)
        # U: Total Tidak Didata (21), V: Total Baru (22), W: Total Selisih (23)
        kel_td = ws.cell(row=r, column=15).value
        kel_b = ws.cell(row=r, column=16).value
        kel_sel = ws.cell(row=r, column=17).value
        
        ush_td = ws.cell(row=r, column=18).value
        ush_b = ws.cell(row=r, column=19).value
        ush_sel = ws.cell(row=r, column=20).value
        
        tot_td = ws.cell(row=r, column=21).value
        tot_b = ws.cell(row=r, column=22).value
        tot_sel = ws.cell(row=r, column=23).value
        
        # Safe string conversion
        kel_sel_str = str(kel_sel) if kel_sel is not None else "None"
        ush_sel_str = str(ush_sel) if ush_sel is not None else "None"
        tot_sel_str = str(tot_sel) if tot_sel is not None else "None"
        name_str = str(name) if name is not None else "None"
        no_str = str(no) if no is not None else "None"
        
        print(f"Row {r} | Name: {name_str:<30} | OrigNo: {no_str:<3} | KelSel: {kel_sel_str:<4} | UshSel: {ush_sel_str:<4} | TotSel: {tot_sel_str:<4}")

if __name__ == "__main__":
    main()
